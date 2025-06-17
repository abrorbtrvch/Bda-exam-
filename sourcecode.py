from openpyxl import load_workbook
from datetime import datetime
import math

# ================================
# CONFIGURATION SECTION
# ================================
EXCEL_FILE = 'sagatave_eksamenam.xlsx'
SHEET_NAME = 'Lapa_0'
HEADER_ROW_INDEX = 3
ENABLE_LOGGING = False  # Set to True for debug prints

# ================================
# UTILITY FUNCTIONS
# ================================

def is_number(val):
    """Check if value is numeric (int or float)."""
    return isinstance(val, (int, float))

def log(*args):
    """Print if logging is enabled."""
    if ENABLE_LOGGING:
        print(*args)

# ================================
# LOAD EXCEL AND HEADER MAPPING
# ================================

wb = load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET_NAME]

# Read headers from specified row and map them to column indices (1-based)
headers = {
    cell.value.strip(): idx
    for idx, cell in enumerate(ws[HEADER_ROW_INDEX], start=1)
}

log("Header mapping:", headers)

# ================================
# INITIALIZE RESULT VARIABLES
# ================================

count_ain_skaits = 0               # Task 1
count_high_2015 = 0               # Task 2
count_adulienas = 0               # Task 3
laserjet_sum = 0                  # Task 4 (Sum of prices)
laserjet_count = 0                # Task 4 (Count for average)
korporativais_sum = 0             # Task 5 (Sum of totals)

# ================================
# ITERATE THROUGH DATA ROWS
# ================================

# Start reading data after header
for row in ws.iter_rows(min_row=HEADER_ROW_INDEX + 1, values_only=True):

    # Extract cell values using header mapping
    adrese = row[headers['Adrese'] - 1]
    skaits = row[headers['Skaits'] - 1]
    prioritate = row[headers['Prioritāte'] - 1]
    pieg_datums = row[headers['Piegādes datums'] - 1]
    pilseta = row[headers['Pilsēta'] - 1]
    produkts = row[headers['Produkts'] - 1]
    cena = row[headers['Cena'] - 1]
    klients = row[headers['Klients'] - 1]
    kopa = row[headers['Kopā'] - 1]

    # Task 1: Count addresses starting with "Ain" and Skaits < 40
    if adrese and isinstance(skaits, (int, float)) and adrese.startswith("Ain") and skaits < 40:
        count_ain_skaits += 1
        log("Matched Task 1:", adrese, skaits)

    # Task 2: Count where Prioritāte is "High" and year is 2015
    if prioritate == 'High' and isinstance(pieg_datums, datetime) and pieg_datums.year == 2015:
        count_high_2015 += 1
        log("Matched Task 2:", prioritate, pieg_datums)

    # Task 3: Count addresses containing "Adulienas iela" and city is Valmiera or Saulkrasti
    if adrese and 'Adulienas iela' in adrese and pilseta in ['Valmiera', 'Saulkrasti']:
        count_adulienas += 1
        log("Matched Task 3:", adrese, pilseta)

    # Task 4: Average Cena of LaserJet products
    if produkts and 'LaserJet' in produkts and is_number(cena):
        laserjet_sum += cena
        laserjet_count += 1
        log("Matched Task 4:", produkts, cena)

    # Task 5: Total Kopā for "Korporatīvais" clients with Skaits between 40–50
    if klients == 'Korporatīvais' and is_number(skaits) and 40 <= skaits <= 50 and is_number(kopa):
        korporativais_sum += kopa
        log("Matched Task 5:", klients, skaits, kopa)

# ================================
# POST-PROCESSING CALCULATIONS
# ================================

average_laserjet = math.floor(laserjet_sum / laserjet_count) if laserjet_count else 0
total_korporativais = math.floor(korporativais_sum)

# ================================
# FINAL OUTPUT REPORT
# ================================

print("\n📊 FINAL REPORT (All Tasks Calculated)")
print("--------------------------------------------------")
print(f"1. Count of 'Ain' addresses with Skaits < 40         : {count_ain_skaits}")
print(f"2. Count of 'High' priority deliveries in year 2015  : {count_high_2015}")
print(f"3. Entries with 'Adulienas iela' in Valmiera/Saulkrasti: {count_adulienas}")
print(f"4. Avg Cena for 'LaserJet' products (rounded down)   : {average_laserjet}")
print(f"5. Total Kopā for 'Korporatīvais' (Skaits 40–50)     : {total_korporativais}")
print("--------------------------------------------------")
print("✅ All tasks completed successfully.\n")
